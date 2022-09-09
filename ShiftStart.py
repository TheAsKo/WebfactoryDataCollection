##### Excel Auto Creation depending on machine
###############################################
# Imports
import logging
import os
import time
import math
import openpyxl
from copy import copy
import mouse
import keyboard
from CurrentMainV2 import WindowFullScreen
###############################################
# Declarations
logging.getLogger().setLevel(logging.DEBUG)
global TableIndex
TableIndex={"Sheet":[],"Index":[],"Shift":[],"Aligment":[],} 
## Copy a sheet with style, format, layout, ect. from one Excel file to another Excel file
## TY StackOverflow

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
        #print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
###############################################
###### Template.xlsx -> AutoData {Machine(x)-8H or 12H)
######               -> Sledovanie{Machine(x)-8/12H Table+Final Doc 

def is_even(num):
    return num % 2 == 0

def is_odd(num):
    return num % 2 != 0

class FileCreation():
    def __run__(AutoDataDirtyInput,EditFileDirtyInput,MachineDict,PathName):
        FileCreation.SheetData=MachineDict
        FileCreation.AutoData(AutoDataDirtyInput)
        ########################################################
        os.startfile('AutoData_gen.xlsx') #NOT PRETTY WAY / NEED TO RESAVE FILE IN EXCEL , FILE HAS BROKEN THEMES ON CREATION SOMEHOW
        time.sleep(2)
        WindowFullScreen()
        time.sleep(1)
        MouseCur=mouse.get_position() #Close Window (os.kill dont work without admin privileges)
        keyboard.press_and_release('Ctrl+S')
        mouse.move(1900,10)
        mouse.click("left")
        mouse.move(MouseCur[0],MouseCur[1]) 
        ########################################################       
        FileCreation.EditFile(EditFileDirtyInput,PathName)
    
    def AutoData(DirtyInput):
        try:
            os.remove('AutoData_gen.xlsx')
        except:
            logging.debug('File deletion failed..')
        else:
            logging.debug('Excel cleaning successfull..')
        if DirtyInput==0 :
            wb_target = openpyxl.Workbook() ####AUTODATA
            wb_source = openpyxl.load_workbook('Template.xlsx')
            for i in range(len(FileCreation.SheetData['Machine'])):
                print(i)
                target_sheet = wb_target.create_sheet('AutoData')
                source_sheet = wb_source['AutoData-Template']
                copy_sheet(source_sheet, target_sheet)
                target_sheet.title='AutoData-'+FileCreation.SheetData['Machine'][i]

            if 'Sheet' in wb_target.sheetnames:  # remove default sheet
                wb_target.remove(wb_target['Sheet'])
                g_sheet=wb_target.sheetnames
            
            for i in range(len(g_sheet)):
                match time.strftime("%H",time.localtime()): #
                    case "06" if FileCreation.SheetData['ShiftCheck'][i]==8  : MachineCheck='Ranná'
                    case "14" if FileCreation.SheetData['ShiftCheck'][i]==8  : MachineCheck='Poobedná'
                    case "22" if FileCreation.SheetData['ShiftCheck'][i]==8  : MachineCheck='Nočná'
                    case "06" if FileCreation.SheetData['ShiftCheck'][i]==12 : MachineCheck='Ranná'
                    case "18" if FileCreation.SheetData['ShiftCheck'][i]==12 : MachineCheck='Nočná'
                    case _ : MachineCheck='ERROR'
                TableIndex['Aligment'] = TableIndex['Aligment'] + [MachineCheck] + [MachineCheck]
                logging.debug('MachineShift: '+str(MachineCheck))
                sheet=wb_target[g_sheet[i]]
                sheet['O2']=str(MachineCheck)
                sheet['P2']=time.strftime("%d.%m.%Y",time.localtime())

            wb_target.save('AutoData_gen.xlsx') #####AUTODATA
            wb_target.close()
            wb_source.close()
            print(TableIndex)  
        else : pass

    def EditFile(DirtyInput,pathname):
        wb_target = openpyxl.Workbook() #####SLEDOVANIE
        wb_source = openpyxl.load_workbook('Template.xlsx')

        target_sheet1 = wb_target.create_sheet("Data")
        source_sheet = wb_source['Data-Template']
        copy_sheet(source_sheet, target_sheet1)
        target_sheet1.title='Data'

        for i in range(len(FileCreation.SheetData['Machine'])): 
            if DirtyInput==1:
                target_sheet = wb_target.create_sheet('AutoData') #DirtyInput
                source_sheet = wb_source['AutoData-Template']
                copy_sheet(source_sheet, target_sheet)
                target_sheet.title='AutoData-'+str(FileCreation.SheetData['Machine'][i])
            else : pass

            target_sheet = wb_target.create_sheet(str(FileCreation.SheetData['ShiftCheck'][i])) #Editable Table
            source_sheet = wb_source[str(FileCreation.SheetData['ShiftCheck'][i])+'H-Template']
            copy_sheet(source_sheet, target_sheet)
            target_sheet.title=FileCreation.SheetData['Machine'][i]+'_'+str(FileCreation.SheetData['ShiftCheck'][i])+'H'

            target_sheet2 = wb_target.create_sheet('Final'+str(FileCreation.SheetData['ShiftCheck'][i])+'H') #Final Doc
            source_sheet = wb_source['Final'+str(FileCreation.SheetData['ShiftCheck'][i])+'H-Template']
            copy_sheet(source_sheet, target_sheet2)
            target_sheet2.title='Hodinové sledovanie-'+str(FileCreation.SheetData['Machine'][i])

        if 'Sheet' in wb_target.sheetnames:  # remove default sheet
            wb_target.remove(wb_target['Sheet'])

        g_sheet=wb_target.sheetnames
        logging.debug(g_sheet)
        print(FileCreation.SheetData)

        for i in range(len(g_sheet)): #NOT PRETTY INDEXING METHOD
            logging.debug(g_sheet[i])
            print(i)
            print(math.floor(i/2)) 
            match g_sheet[i][0:2]: 
                case 'Da' : logging.debug('Skipping editing Data') ; continue #Another skipping of Data Sheet 
                case 'Ho' : logging.debug('FinalTable') ; IndexCase=1      
                case _    : logging.debug('EditingTable') ; IndexCase=2 #NOT SAFE DEFAULTING
            TableIndex['Sheet'] = TableIndex['Sheet'] + [g_sheet[i]]
            TableIndex['Index'] = TableIndex['Index'] + [IndexCase]
            print(TableIndex)
            x = is_odd(i) #NOT PRETTY AT ALL BUT I CANT THINK OF ANYTHING BETTER
            match i: #i0-2=x0,i3-4=x1,i5-6=x2,i7-8=x3,i9-10=x4,....
                case 0 | 1 | 2 : x=0
                case _ : 
                    match x:
                        case True: x=math.floor(i/2)
                        case False: pass

            TableIndex['Shift'] = TableIndex['Shift'] + [FileCreation.SheetData['ShiftCheck'][x]]

        for i in range(len(TableIndex['Sheet'])):
            if TableIndex['Index'][i]==0 : logging.debug('DataSheet : '+TableIndex['Sheet'][i]) #Always Skip Data Sheet
            elif TableIndex['Index'][i]==1 :
                logging.debug("Hodinove Sledovanie : "+TableIndex['Sheet'][i])
                CelNum=5 #Formating offset
                CelMax=TableIndex['Shift'][i]+CelNum
                while CelNum<CelMax:
                    sheet=wb_target[TableIndex['Sheet'][i]]
                    x="'"+str(g_sheet[TableIndex['Index'][i]])+"'"
                    sheet['C'+str(CelNum)]="=IF("+x+"!B"+str(CelNum+1)+'="","",'+str(x)+"!B"+str(CelNum+1)+")"
                    sheet['D'+str(CelNum)]="=IF("+x+"!C"+str(CelNum+1)+'="","",'+str(x)+"!C"+str(CelNum+1)+")"
                    sheet['E'+str(CelNum)]="=IF("+x+"!D"+str(CelNum+1)+'="","",'+str(x)+"!D"+str(CelNum+1)+")"
                    sheet['F'+str(CelNum)]="=IF("+x+"!E"+str(CelNum+1)+'="","",'+str(x)+"!E"+str(CelNum+1)+")"
                    sheet['G'+str(CelNum)]="=IF("+x+"!F"+str(CelNum+1)+'="","",'+str(x)+"!F"+str(CelNum+1)+")"
                    sheet['I'+str(CelNum)]="=IF("+x+"!T"+str(CelNum+1)+'="","",'+str(x)+"!T"+str(CelNum+1)+")"
                    CelNum=CelNum+1
                sheet['I2']='Pracovisko: '+str(TableIndex['Sheet'][i][20:])   
                sheet['J2']='="Dátum: "&'+str(x)+'!G3'
                #sheet['I3']='Meno+os.číslo:'
                sheet['J3']='="Zmena: "&'+str(x)+'!C3'
                        
            elif TableIndex['Index'][i]==2 :
                CelNum=6 #Formating offset
                CelMax=TableIndex['Shift'][i]+CelNum
                logging.debug("ShiftTable : "+TableIndex['Sheet'][i])
                if   TableIndex['Shift'][i]==8 : y="Data!$B$3:$J$6"  #=VLOOKUP($C$3;Data!$B$3:$J$6;2)
                elif TableIndex['Shift'][i]==12: y="Data!$B$9:$N$11"  #=VLOOKUP($C$3;Data!$B$9:$N$11;2)
                else : y="ERROR" ; logging.critical("Failed to allocate hour table!")
                x="'"+pathname+'\[AutoData_gen.xlsx]AutoData-'+TableIndex['Sheet'][i]+"'" #Pathname+filename+sheet
                while CelNum<CelMax:
                    OKCell = "=IF(I"+str(CelNum)+">0,I"+str(CelNum)+",IF("+x+"!D"+str(CelNum-4)+'="","",'+str(x)+"!D"+str(CelNum-4)+"))"
                    NOKCell = "=IF(J"+str(CelNum)+">0,J"+str(CelNum)+",IF("+x+"!E"+str(CelNum-4)+'="","",'+str(x)+"!E"+str(CelNum-4)+"))" 
                    sheet=wb_target[TableIndex['Sheet'][i]] #VAR1 8H R-P-N 
                    if TableIndex['Shift'][i] == 12 and TableIndex['Aligment'][i] == 'Nočná': #VAR 3 12H Nocna
                        match CelNum:
                            case 6 | 7 | 8 | 9 : pass #x+Pre0
                            case 10 | 11 | 12 | 13 | 14 | 15 | 16 | 17 | 18 : pass #x+21
                    elif TableIndex['Shift'][i] == 12 and TableIndex['Aligment'][i] == 'Ranná': # VAR2 12H R
                        match CelNum :
                            case 14 | 15 | 16 | 17 | 18 : OKCell = OKCell[:-2] + 'G13))' ; NOKCell = NOKCell[:-2] + 'H13))'
                            case _ : pass 
                    sheet['A'+str(CelNum)]="=IF("+x+"!B"+str(CelNum-4)+'="","",'+str(x)+"!B"+str(CelNum-4)+")"
                    sheet['B'+str(CelNum)]='=VLOOKUP($C$3,'+y+','+str(CelNum-4)+')'
                    sheet['C'+str(CelNum)]="=IF("+x+"!F"+str(CelNum-4)+'="","",'+str(x)+"!F"+str(CelNum-4)+")"
                    sheet['D'+str(CelNum)]="=IF("+x+"!H"+str(CelNum-4)+'="","",'+str(x)+"!H"+str(CelNum-4)+")"
                    sheet['G'+str(CelNum)]=OKCell
                    sheet['H'+str(CelNum)]=NOKCell
                    sheet['P'+str(CelNum)]="=IF("+x+"!G"+str(CelNum-4)+'="","",'+str(x)+"!G"+str(CelNum-4)+")"
                    sheet['Q'+str(CelNum)]="=IF("+x+"!C"+str(CelNum-4)+'="","",'+str(x)+"!C"+str(CelNum-4)+")"
                    CelNum=CelNum+1
                sheet['C3']='='+x+'!O2'
                sheet['G3']='='+x+'!P2'

        wb_target.save('Hodinové_sledovanie_gen.xlsx') #####Hodinove
        wb_target.close()
        wb_source.close()

#########################################
