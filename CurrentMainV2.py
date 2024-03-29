#Main Code - Image Processing + Excel Data Output
###############################################
# Imports
from numpy import negative
import pyautogui
import time
import os
from PIL import Image
import pytesseract
from openpyxl import load_workbook
import mouse
import logging
import cv2
import pygetwindow as gw
##################################
# Declarations
SheetDict = {"TimeFlag":["A2","1"],"TimeSend":["B2",'1'],'OEE':['C2','1'],'OK':['D2','1'],'NOK':['E2','1'],'Product':['F2','1'],"Scrap":["G2",'1'],"Norm":["H2","1"]}
ScreenRegionDict={"OEE":(600,230,210,70),'OK':(1300,680,260,100),'NOK':(1680,680,190,80),"Product":(1020,620,600,50),"Scrap":(1700,230,120,40),"Norm":(1020,670,120,50)}
DEBUG=1 # For my personal use , maybe cleaned up after ....
OnlyRootDebug=True #Used to disable all libs from using logging :)
pytesseract.pytesseract.tesseract_cmd = 'C:/Users/nsz.fu.montaz/AppData/Local/Tesseract-OCR/tesseract.exe' #Tesseract .exe bcs i cant use PATH on admin locked work device :/ need to add to GUI + maybe auto location
tessdefault_config = "-c tessedit_char_whitelist 0123456789., tessedit_char_blacklist _"
tessnumber_config='--psm 3 --oem 3 -c tessedit_char_whitelist=0123456789/-.%, tessedit_char_blacklist _'
logging.getLogger().setLevel(logging.DEBUG) #Need to add to GUI as check + need to clean/add logging
URLDict=('https://wf-nsm.neuman.at/clients/wf-login/#/','http://wf-nsm.neuman.at/clients/wf-mes/sk/#/wfmes/view/(mainview:msc/349)','http://wf-nsm.neuman.at/clients/wf-mes/sk/#/wfmes/view/(mainview:msc/346)')
#URLDict #Need to add some sort of autologin at start of the shift
DeleteImagesAfterUsage=0 #Auto-cleaning of all images , maybe i should add all unnecessary files (need to be true for sure when release)
###############################################
# Classes and Definitions
def SheetDictData(x,ShiftCheck): #Assign proper cell number depending on hour of data collection
        match x:        #Maybe move to time table in time loop ?
            case 7 | 19 if ShiftCheck == 12: return 2;
            case 8 | 20 if ShiftCheck == 12: return 3;
            case 9 | 21 if ShiftCheck == 12: return 4;
            case 10| 22 if ShiftCheck == 12: return 5;
            case 11| 23 if ShiftCheck == 12: return 6;
            case 12| 24 if ShiftCheck == 12: return 7;
            case 13| 1  if ShiftCheck == 12: return 8;
            case 14| 2  if ShiftCheck == 12: return 9;
            case 15| 3  if ShiftCheck == 12: return 10;
            case 16| 4  if ShiftCheck == 12: return 11;
            case 17| 5  if ShiftCheck == 12: return 12;
            case 18| 6  if ShiftCheck == 12: return 13;
            case 7 | 15 | 23 if ShiftCheck == 8: return 2;
            case 8 | 16 | 24 if ShiftCheck == 8: return 3;
            case 9 | 17 | 1  if ShiftCheck == 8: return 4;
            case 10| 18 | 2  if ShiftCheck == 8: return 5;
            case 11| 19 | 3  if ShiftCheck == 8: return 6;
            case 12| 20 | 4  if ShiftCheck == 8: return 7;
            case 13| 21 | 5  if ShiftCheck == 8: return 8;
            case 14| 22 | 6  if ShiftCheck == 8: return 9;
            case _ : logging.critical("Failed to allocate data to sheet disc");return 0,    

def ScreenshotRegion(name,Left,Top,Width,Height): #Making screenshot , not really needed but code looks cleaner
    logging.debug("Creating screenshot : "+name+".png")
    pyautogui.screenshot(str(name+".png"),region=(Left,Top,Width,Height))

def LoadCheck(): #Checking if webpage is loaded fully...
    ScreenshotRegion("LoadCheck",0,980,180,40)
    ScreenshotRegion("ScrapLoadCheck",1820,230,70,70) #Using 2 separate ways of comparing images is not best but LoadCheck wasnt working with cv2.norm somehow
    b1, g1, r1 = cv2.split(cv2.subtract(cv2.imread("LoadCheck.png"),cv2.imread("LoadGood.png"))) # This is good only if both things are vibrant colors
    ImgDiffGreen = 1 - cv2.norm(cv2.imread("ScrapLoadCheck.png"),cv2.imread("ScrapLoadGoodGreen.png"), cv2.NORM_L2 ) / ( 70 * 70 )
    ImgDiffRed = 1 - cv2.norm(cv2.imread("ScrapLoadCheck.png"),cv2.imread("ScrapLoadGoodRed.png"), cv2.NORM_L2 ) / ( 70 * 70 )
    logging.debug('ImgDiffGreen similarity = '+str(ImgDiffGreen)) 
    logging.debug('ImgDiffRed similarity = '+str(ImgDiffRed))
    if cv2.countNonZero(b1) == 0  and cv2.countNonZero(g1) == 0 and cv2.countNonZero(r1) == 0 and ImgDiffGreen>0.40 or ImgDiffRed>0.40: #0.40 is for moused grayness , 0.9 is normal
        if DeleteImagesAfterUsage==1: #Extremely lazy cleaning ....
            try:
                os.remove("LoadCheck.png")
                os.remove("ScrapLoadCheck.png")
            except:
                pass
            else:
                pass
        logging.info("Loading of page completed...")
    else:
        logging.info('Loading not finished , waiting...')
        return 0; #replacing semicolon breaks whole cycling :)

def ImageProcess(IMGName,range2=-1,range=0,tescfg=tessnumber_config): #Image Data Extraction
    try:
        img = cv2.imread(str(IMGName)+'.png')
        img = cv2.resize(img, None, fx=2, fy=2,interpolation=cv2.INTER_CUBIC)
        value=pytesseract.image_to_string(img,config=tescfg)[range:range2]    
    except:
        logging.debug(IMGName+" error!")
        if DEBUG==1 :
            logging.exception('')
        return "error"
    else:
        if DEBUG==1:logging.debug(IMGName+" processed...")
        return value

#print ([k for k in logging.Logger.manager.loggerDict]) #### totally not from stackoverflow :)

def ImageGrab(MachineName,HourValue,MachineURL):
    logging.getLogger('PIL.Image').disabled = OnlyRootDebug
    logging.getLogger('PIL').disabled = OnlyRootDebug
    logging.getLogger('PIL.TiffImagePlugin').disabled = OnlyRootDebug
    logging.getLogger('PIL.PngImagePlugin').disabled = OnlyRootDebug ### Some loggers get loaded during using of sublibs ....

    os.system("start chrome --start-maximized --new-window "+URLDict[MachineURL]) #Start Browser (need to add auto login - most likely just mouse clicks)
    
    time.sleep(1) #Maybe not needed , i just want to be sure to get proper window selection
    window=gw.getActiveWindow()   #Forcing window maximizing
    if window.isMaximized == True: 
        logging.debug("Window is maximized")
    else :
        window.maximize()
        logging.debug("Maximizing window AGAIN!") 

    while LoadCheck()==0: #not great way to check i think but work
        time.sleep(1)
        logging.info("Waiting for page load")

    for each in ScreenRegionDict:
        logging.debug("Image Creation: "+each)
        x=ScreenRegionDict[each] #Redundant line but looks better ? + less listing through dict ? performance gains ?
        ScreenshotRegion(each+MachineName+str(HourValue),x[0],x[1],x[2],x[3])

    MouseCur=mouse.get_position() #Close Window (not pretty simualing mouse clicks but i dont want to mess with os.kill and os.pid)
    mouse.move(10,10)
    mouse.click("middle")
    mouse.move(MouseCur[0],MouseCur[1])

def ExcelOutput(MachineName,HourValue,ShiftCheck):
    logging.info("Applying data to excel")  #Excel Start
    #wb = load_workbook(filename = str(ExcelFile[ExcelMain]))
    wb = load_workbook(filename="AutoData.xlsx")
    sheet=wb["AutoData-"+MachineName]
    
    # !!MESSY!!
    SheetDictDataAnswer=SheetDictData(int(HourValue),ShiftCheck) #Getting info from main thread
    for each in SheetDict: #Editing cell location 
        x = SheetDict[each][0][:-1]
        x = x+str(SheetDictDataAnswer)
        SheetDict[each][0] = str(x)
        match each: #Editing cell value
            case "TimeFlag" : SheetDict[each][1] = str(SheetDictDataAnswer-1) #-1 = Offset used in excel for header
            case "TimeSend" : SheetDict[each][1] = time.strftime("%H:%M:%S",time.localtime())
            case "OEE" | "Scrap" : SheetDict[each][1] = ImageProcess(each+MachineName+str(HourValue),-2) #Extra removing % , it safer to detect and then remove than to block with blacklist
            case "Product" : SheetDict[each][1] = ImageProcess(each+MachineName+str(HourValue),-5,tescfg='--psm 7 --oem 3') #Should be good , need more data
            case "OK" : #Hardcoded cleaning of value + Maybe still broken + Not pretty at all
                x = ImageProcess(each+MachineName+str(HourValue),4)
                if x=="error": #Not pretty handling
                    SheetDict[each][1] = x
                    break  
                while x.isdigit() == False : #Still too many infinite loops
                    logging.debug("OK before:"+x)
                    y=negative(len(x)-1) 
                    #x = x[y:]
                    x = x[:-y]
                    logging.debug("OK after :"+x)
                    time.sleep(0.5)   #Maybe still broken , need to log data
                SheetDict[each][1] = x
            case 'NOK': # NEED TO FINISH , BAD REPORTING OF 0,1 
                SheetDict[each][1] = ImageProcess(each+MachineName+str(HourValue),-1,tescfg="--psm 7 --oem 3") #Need to log data
            case _ : SheetDict[each][1] = ImageProcess(each+MachineName+str(HourValue),-1)

        logging.debug(str(each)+" : "+str(SheetDict[each][0]+' : '+str(SheetDict[each][1])))
        sheet[str(SheetDict[each][0])]=str(SheetDict[each][1])   #Writing into excel
        if DeleteImagesAfterUsage==1: #Extremely lazy cleaning ....
            try:
                os.remove(str(each+MachineName+str(HourValue))+".png")
            except:
                pass
            else:
                pass
    # !!MESSY!!

    try: #Saving Excel Document #Very barebones 
        #wb.save(str(ExcelFile[ExcelMain]))
        wb.save(filename="AutoData.xlsx")  
    except:
        logging.warning("Excel edit failed") 
    else:
        logging.info('Excel edit success') 




